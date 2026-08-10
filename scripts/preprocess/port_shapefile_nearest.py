#!/usr/bin/env python
# coding: utf-8

import os
import re
import geopandas as gpd
import pandas as pd

from utils_new import load_config


EXTRA_ASIA_ISO3 = {"RUS", "TUR", "GEO", "ARM", "AZE", "CYP"} #ds extra countires that might not come under Asia or Pacific
REGION_COUNTRY_ALLOWLIST = {
    "northern_russia": {"RUS"},
    "pacific": None,
}

COUNTRY_ALIASES = { # alternative names for countries, abd there iso standard name
    "south korea": "Korea, Republic of",
    "korea republic of": "Korea, Republic of",
    "korea": "Korea, Republic of",
    "republic of korea": "Korea, Republic of",
    "south korea republic of": "Korea, Republic of",
    "north korea": "Korea (Democratic People's Republic of)",
    "korea democratic people republic of": "Korea (Democratic People's Republic of)",
    "democratic peoples republic of korea": "Korea (Democratic People's Republic of)",
    "north korea democratic peoples republic of": "Korea (Democratic People's Republic of)",
    "russia": "Russian Federation",
    "russian federation": "Russian Federation",
    "ussr": "Russian Federation",
    "soviet union": "Russian Federation",
    "viet nam": "Viet Nam",
    "vietnam": "Viet Nam",
    "turkey": "Türkiye",
    "turkiye": "Türkiye",
    "turkish republic": "Türkiye",
    "türkiye": "Türkiye",
    "turk republic": "Türkiye",
    "turkey republic": "Türkiye",
    "taiwan": "Taiwan, Province of China",
    "taiwan province of china": "Taiwan, Province of China",
    "hong kong": "Hong Kong",
    "macao": "Macao",
    "macau": "Macao",
    "brunei": "Brunei Darussalam",
    "brunei darussalam": "Brunei Darussalam",
    "east timor": "Timor-Leste",
    "timor leste": "Timor-Leste",
    "palestine": "Palestine, State of",
    "palestine state of": "Palestine, State of",
    "iran": "Iran (Islamic Republic of)",
    "iran islamic republic of": "Iran (Islamic Republic of)",
    "micronesia": "Micronesia (Federated States of)",
    "micronesia federated states of": "Micronesia (Federated States of)",
    "syrian arab republic": "Syrian Arab Republic",
    "the netherlands": "Netherlands",
    "netherlands": "Netherlands",
    "holland": "Netherlands",
    "dutch republic": "Netherlands",
    "united kingdom": "United Kingdom",
    "uk": "United Kingdom",
    "great britain": "United Kingdom",
    "britain": "United Kingdom",
    "england": "United Kingdom",
    "scotland": "United Kingdom",
    "wales": "United Kingdom",
    "northern ireland": "United Kingdom",
}


def normalize_country(value):
    """Normalize a country name or ISO token for robust matching."""
    if pd.isna(value): # ds checked wheter value is missing
        return ""

    value = str(value).strip().lower() # ds converts to lowercase and removes spaces
    value = re.sub(r"[^a-z0-9\s]", "", value) #ds removes all symbols 
    value = re.sub(r"\s+", " ", value) #ds collapse multiple spaces into one space
    return value.strip() #ds removes spaces from beginning and end


def load_country_iso_allow_list(countries_xlsx):
    """Load the country workbook and return the ISO3 allow-list and name lookup."""
    if not os.path.exists(countries_xlsx): #ds check excel sheet exists there
        raise FileNotFoundError(f"Countries workbook not found: {countries_xlsx}")

    ref_df = pd.read_excel(countries_xlsx) 
    ref_df.columns = [str(col).strip() for col in ref_df.columns] # removes extra spaces and turns them into strings for the column names

    required_columns = {"Country/Territory", "iso2", "iso3"} 
    missing = required_columns - set(ref_df.columns)
    if missing:
        raise ValueError(
            f"Countries workbook is missing required columns: {', '.join(sorted(missing))}"
        )

    ref_df = ref_df[["Country/Territory", "iso2", "iso3"]].rename(
        columns={"Country/Territory": "country_name"} # renames first column name to be compatiable in python
    )
    #reads columns
    ref_df["country_name"] = ref_df["country_name"].astype(str).str.strip()
    ref_df["iso2"] = ref_df["iso2"].fillna("").astype(str).str.strip().str.upper()
    ref_df["iso3"] = ref_df["iso3"].fillna("").astype(str).str.strip().str.upper()

    allowed_iso3 = { # turns iso into a list
        iso for iso in ref_df["iso3"].tolist()
        if iso and iso.lower() != "nan"
    }

    country_lookup = {} # goes through every row
    for _, row in ref_df.iterrows():
        key = normalize_country(row["country_name"])
        iso3 = row["iso3"]
        if not key or not iso3 or iso3.lower() == "nan": # checks that none of the entries are nan
            continue
        if key in country_lookup and country_lookup[key] != iso3:  #ds Check that duplicate country names have the same ISO3 code.
            raise ValueError(
                f"Duplicate country name after normalisation: {row['country_name']}"
            )
        country_lookup[key] = iso3 # matches the county and iso in a dictonary 

    for alias, canonical in COUNTRY_ALIASES.items(): # ds checks alternative country names
        canonical_norm = normalize_country(canonical)
        if canonical_norm in country_lookup:
            country_lookup[normalize_country(alias)] = country_lookup[canonical_norm] # replace all the names so they are now iso3 standard named

    return ref_df, country_lookup, allowed_iso3 # ds ref_df = adapted country reference table, country_lookup = dictionary mapping normalised country names to ISO3 codes, allowed_iso3 = set of all valid ISO3 codes in the workbook


def country_to_iso3(country_value, country_lookup):
    """Map a landuse country string to the workbook ISO3 code when possible."""
    if pd.isna(country_value): # ds if there is a country value missing in landuse return a panda missing value
        return pd.NA

    value = normalize_country(country_value) # ds checks after it is cleaned if there are any missing values (catches any "", first if isnt necessary)
    if not value: 
        return pd.NA

    return country_lookup.get(value, pd.NA) # ds searches for the corresponding iso3 value for the value in country_lookup, if not returns NA 


def apply_region_country_filter(landuse_gdf, region_name, allowed_iso3):
    """Keep only the ISO3 countries allowed for the study area
    """
    if region_name is None:
        return landuse_gdf.loc[landuse_gdf["country_iso3"].isin(allowed_iso3)].copy() # ds if no region is specified, return all landuse records that are in the allowed_iso3 set

    region_allowed = REGION_COUNTRY_ALLOWLIST.get(region_name) # ds is the region in the allowed list
    if region_allowed is None:
        return landuse_gdf.loc[landuse_gdf["country_iso3"].isin(allowed_iso3)].copy()

    keep_iso3 = set(allowed_iso3).intersection(region_allowed) # keep only the iso3 in the allowed region
    return landuse_gdf.loc[landuse_gdf["country_iso3"].isin(keep_iso3)].copy()


def build_nearest_port_lookup(landuse_gdf, network_ports):
    """Use nearest-point matching against the full processed network port node layer."""
    # Only keep port nodes for the nearest-port lookup.
    network_ports = network_ports.loc[network_ports["infra"] == "port"].copy() # ds extract the ports (exlcudes maritime nodes)

    # Reproject both layers to a metric CRS so the nearest-distance calculation is meaningful.
    projected_landuse = landuse_gdf.to_crs("EPSG:3857")
    projected_network = network_ports.to_crs("EPSG:3857")

    candidate_ports = projected_network[
        ["id", "name", "country", "iso3", "geometry"]
    ].rename(
        columns={
            "id": "nearest_port_id",
            "name": "nearest_port_name",
            "country": "nearest_port_country",
            "iso3": "nearest_port_iso3",
        }
    )

    nearest = gpd.sjoin_nearest( # ds find nearest node
        projected_landuse,
        candidate_ports,
        how="left",
        distance_col="distance_m",
    )

    # Restore the original CRS on the output layer for writing back to disk.
    nearest = nearest.set_geometry("geometry")
    nearest = nearest.to_crs(landuse_gdf.crs)

    return nearest


def main(config):
    incoming_data_path = config["paths"]["incoming_data"]
    processed_data_path = config["paths"]["processed_data"]

    landuse_gpkg = os.path.join(incoming_data_path, "infrastructure", "Port", "port_landuse.gpkg")
    countries_xlsx = os.path.join(incoming_data_path, "Countries_list.xlsx")
    network_gpkg = os.path.join(
    processed_data_path,
    "infrastructure",
    "port",
    "asia_pacific_maritime_network_PROVA_NEW1.gpkg"
)

    output_dir = os.path.join(processed_data_path, "infrastructure","port")
    os.makedirs(output_dir, exist_ok=True) #ds create directories if they do not exist already

    landuse_gdf = gpd.read_file(landuse_gpkg)
    network_ports = gpd.read_file(network_gpkg, layer="port_nodes")

    # Load country ISO reference list.
    country_ref, country_lookup, allowed_iso3 = load_country_iso_allow_list(countries_xlsx)
    allowed_iso3 = allowed_iso3.union(EXTRA_ASIA_ISO3) # ds assure all iso3 in strudy area are included

    # Map landuse country names into the workbook ISO3 set first.
    landuse_gdf["country_iso3"] = landuse_gdf["country"].apply(lambda x: country_to_iso3(x, country_lookup))
    landuse_gdf["country_in_allowed_list"] = landuse_gdf["country_iso3"].isin(allowed_iso3)

    # Filter the landuse records to the countries that appear in the workbook before any nearest match.
    landuse_gdf = landuse_gdf.loc[landuse_gdf["country_in_allowed_list"]].copy()

    # ds Filter to only include relevant countries in the study area
    landuse_gdf = apply_region_country_filter(landuse_gdf, region_name=None, allowed_iso3=allowed_iso3)

    # Normalize the landuse country labels for reporting.
    landuse_gdf["country_norm"] = landuse_gdf["country"].apply(normalize_country)

    # Run the nearest-port lookup against the full network port node set.
    nearest = build_nearest_port_lookup(landuse_gdf, network_ports)

    # Keep a tidy export that contains every landuse row and its nearest port id.
    lookup_export = nearest[
        [
            "port_name", #from land_use
            "country", #from land use
            "continent", #from land use
            "area", #from land use
            "type", #from land_use
            "sector", #from land_use
            "land_use", #from land_use
            "nearest_port_id", # from maritime_nerwork_new1.gpkg
            "nearest_port_name", # from maritime_nerwork_new1.gpkg  
            "nearest_port_country", 
            "nearest_port_iso3", 
            "distance_m", # calculated, distance between port and node
        ]
    ].copy()

    output_landuse = os.path.join(output_dir, "port_landuse_nearest_port_ids.gpkg")
    nearest.to_file(output_landuse, layer="port_landuse", driver="GPKG")

    # Build a node polygon audit sheet.
    # If your polygon layer uses a different linking field, change polygon_node_id_col above.
    polygon_counts = (
        nearest.groupby("nearest_port_id")
        .size()
        .reset_index(name="polygon_count")
    )

    node_polygon_audit = network_ports[
        ["id", "name", "infra", "country", "iso3"]
    ].copy()

    node_polygon_audit = node_polygon_audit.merge(
        polygon_counts,
        left_on="id",
        right_on="nearest_port_id",
        how="left"
    )

    node_polygon_audit["polygon_count"] = node_polygon_audit["polygon_count"].fillna(0).astype(int)
    node_polygon_audit["missing_polygon_data"] = node_polygon_audit["polygon_count"] == 0
    node_polygon_audit.drop(columns=["nearest_port_id"], inplace=True, errors="ignore")

    # Save everything into one Excel workbook with separate sheets.
    lookup_xlsx = os.path.join(output_dir, "port_landuse_nearest_port_id_lookup.xlsx")

    summary_sheet = pd.DataFrame({
        "metric": [
            "total_records",
            "matched_records",
            "unmatched_records",
            "nodes_total",
            "nodes_missing_polygon_data",
            "nodes_with_polygons",
        ],
        "value": [
            len(lookup_export),
            int(lookup_export["nearest_port_id"].notna().sum()),
            int(lookup_export["nearest_port_id"].isna().sum()),
            len(node_polygon_audit),
            int(node_polygon_audit["missing_polygon_data"].sum()),
            int((node_polygon_audit["polygon_count"] > 0).sum()),
        ]
    })

    with pd.ExcelWriter(lookup_xlsx, engine="openpyxl") as writer:
        lookup_export.to_excel(writer, sheet_name="lookup", index=False)
        node_polygon_audit.to_excel(writer, sheet_name="node_polygon_audit", index=False)
        summary_sheet.to_excel(writer, sheet_name="summary", index=False)

        ws = writer.book["lookup"]

        from openpyxl.styles import PatternFill

        green_fill = PatternFill(fill_type="solid", fgColor="C6EFCE")
        red_fill = PatternFill(fill_type="solid", fgColor="FFC7CE")

        header_map = {cell.value: cell.column for cell in ws[1]}
        port_col = header_map["port_name"]
        nearest_col = header_map["nearest_port_name"]

        for row in range(2, ws.max_row + 1):
            port_value = ws.cell(row=row, column=port_col).value
            nearest_value = ws.cell(row=row, column=nearest_col).value

            port_value = "" if port_value is None else str(port_value).strip()
            nearest_value = "" if nearest_value is None else str(nearest_value).strip()

            cell = ws.cell(row=row, column=nearest_col)

            if port_value == nearest_value:
                cell.fill = green_fill
            else:
                cell.fill = red_fill

        ws.freeze_panes = "A2"
        # Add conditional formatting to the node_polygon_audit sheet.
    audit_ws = writer.book["node_polygon_audit"]

    from openpyxl.styles import PatternFill

    green_fill = PatternFill(fill_type="solid", fgColor="C6EFCE")
    red_fill = PatternFill(fill_type="solid", fgColor="FFC7CE")

    audit_header_map = {cell.value: cell.column for cell in audit_ws[1]}
    missing_col = audit_header_map["missing_polygon_data"]

    for row in range(2, audit_ws.max_row + 1):
        cell = audit_ws.cell(row=row, column=missing_col)

        if cell.value is True:
            cell.fill = red_fill
        else:
            cell.fill = green_fill

    audit_ws.freeze_panes = "A2"
    from openpyxl.styles import PatternFill

    from openpyxl.styles import PatternFill

    with pd.ExcelWriter(lookup_xlsx, engine="openpyxl") as writer:
        lookup_export.to_excel(writer, sheet_name="lookup", index=False)
        node_polygon_audit.to_excel(writer, sheet_name="node_polygon_audit", index=False)
        summary_sheet.to_excel(writer, sheet_name="summary", index=False)

        ws = writer.book["lookup"]

        green_fill = PatternFill(fill_type="solid", fgColor="C6EFCE")
        red_fill = PatternFill(fill_type="solid", fgColor="FFC7CE")

        # Get header positions
        headers = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}
        country_col = headers["country"]
        nearest_country_col = headers["nearest_port_country"]

        # Color the whole row based on whether country matches nearest_port_country
        for row in range(2, ws.max_row + 1):
            country_value = ws.cell(row=row, column=country_col).value
            nearest_country_value = ws.cell(row=row, column=nearest_country_col).value

            country_value = "" if country_value is None else str(country_value).strip()
            nearest_country_value = "" if nearest_country_value is None else str(nearest_country_value).strip()

            fill = green_fill if country_value == nearest_country_value else red_fill

            for col in range(1, ws.max_column + 1):
                ws.cell(row=row, column=col).fill = fill

        ws.freeze_panes = "A2"



    unmatched_count = lookup_export["nearest_port_id"].isna().sum()
    matched_count = len(lookup_export) - unmatched_count

    print(f"Allowed ISO3 countries in workbook: {len(country_ref)}")
    print(f"Nearest-port matches kept: {matched_count}")
    print(f"Nearest-port unmatched records: {unmatched_count}")
    print(f"Saved enriched landuse layer to: {output_landuse}")
    print(f"Saved spreadsheet lookup to: {lookup_xlsx}")

    return nearest, lookup_export, node_polygon_audit


if __name__ == "__main__":
    CONFIG = load_config()
    main(CONFIG)
